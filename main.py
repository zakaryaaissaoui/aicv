import os
import shutil
import hashlib
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Response
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.templating import Jinja2Templates
from io import BytesIO
import tempfile
import qrcode
import qrcode.image.svg
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import our custom modules
import database as db
import parser_engine as parser
import search_engine as search

app = FastAPI(title="TalentFlow AI - API", description="Backend APIs for TalentFlow AI Resume Platform")

# CORS middleware to allow cross-origin requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Setup directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Mount static files (will hold CSS, JS, Images)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

templates = Jinja2Templates(directory=TEMPLATES_DIR)

# Initialize Database on Startup
@app.on_event("startup")
def startup_event():
    db.init_db()
    # Seed initial candidates if table is empty
    candidates = db.get_all_candidates()
    if not candidates:
        print("Database is empty, seeding initial candidates...")
        seed_data = [
            {
                "name": "Hamza Belkaid",
                "phone": "0555 12 34 56",
                "email": "hamza.belkaid@email.com",
                "education": ["Master en Systèmes Mobiles - Université de Tlemcen (2023)"],
                "experience_years": 3.0,
                "experience_details": [
                    "Développeur Flutter Senior - DevApp Tlemcen (2024-Présent)",
                    "Développeur Mobile Junior - AlgerTech (2022-2024)"
                ],
                "skills": ["Flutter", "Dart", "Firebase", "Git", "Android", "iOS", "REST API"],
                "languages": ["Arabic", "French", "English"],
                "wilaya": "Tlemcen",
                "ai_score": 92,
                "ai_summary": "Développeur Flutter talentueux avec 3 ans d'expérience, basé à Tlemcen. Profil solide en développement d'applications mobiles cross-platform et en intégration de Firebase.",
                "strengths": ["Forte maîtrise de Flutter & Dart", "Résident à Tlemcen (proximité locale)", "Expérience de production réelle"],
                "weaknesses": ["Compétences Backend limitées", "Moins d'expérience en DevOps"],
                "pipeline_stage": "Interview",
                "notes": "Très bon entretien technique initial. Compétences Flutter solides.",
                "file_name": "Hamza_Belkaid_Flutter.pdf",
                "file_hash": "mock_hash_hamza_belkaid_123456",
                "file_content_text": "Hamza Belkaid. Email: hamza.belkaid@email.com. Phone: 0555123456. Tlemcen. Flutter Developer with 3 years of experience. Experience: Senior Flutter Developer at DevApp Tlemcen, Junior Mobile Developer at AlgerTech. Education: Master in Mobile Systems from University of Tlemcen. Skills: Flutter, Dart, Firebase, Git, Android, iOS."
            },
            {
                "name": "Amine Meziane",
                "phone": "0661 98 76 54",
                "email": "a.meziane@email.com",
                "education": ["Ingénieur d'État en Informatique - ESI Alger (2021)"],
                "experience_years": 5.0,
                "experience_details": [
                    "Lead Backend Engineer - PySoft Alger (2023-Présent)",
                    "Python Developer - InnovSolutions (2021-2023)"
                ],
                "skills": ["Python", "FastAPI", "PostgreSQL", "Docker", "AWS", "Git", "Redis", "MongoDB"],
                "languages": ["French", "English", "Arabic"],
                "wilaya": "Alger",
                "ai_score": 95,
                "ai_summary": "Lead Backend Python avec 5 ans d'expérience. Profil hautement technique axé sur la conception d'APIs performantes avec FastAPI/Django et le déploiement sur Docker/AWS.",
                "strengths": ["Maîtrise approfondie de Python & FastAPI", "Compétences d'architecture Cloud & DevOps", "Diplômé de l'ESI (École réputée)"],
                "weaknesses": ["Aucune compétence Frontend", "Niveau de salaire attendu élevé"],
                "pipeline_stage": "Test",
                "notes": "A passé le test technique avec un score de 95%. Profil très prometteur.",
                "file_name": "Amine_Meziane_Python_Backend.pdf",
                "file_hash": "mock_hash_amine_meziane_456789",
                "file_content_text": "Amine Meziane. Email: a.meziane@email.com. Phone: 0661987654. Alger. Python Backend Engineer with 5 years experience. Experience: Lead Backend Engineer at PySoft Alger, Python Developer at InnovSolutions. Education: Engineer in Computer Science from ESI Alger. Skills: Python, FastAPI, PostgreSQL, Docker, AWS, Git, Redis, MongoDB."
            },
            {
                "name": "Sara Haddad",
                "phone": "0770 45 67 89",
                "email": "sara.haddad@email.com",
                "education": ["Licence en Informatique - Université d'Oran (2024)"],
                "experience_years": 1.5,
                "experience_details": [
                    "Développeuse Front-end - WebStudio Oran (2024-Présent)",
                    "Stagiaire Intégratrice Web - OranApp (2023)"
                ],
                "skills": ["React", "JavaScript", "HTML", "CSS", "Tailwind CSS", "Git", "Figma"],
                "languages": ["Arabic", "French", "English"],
                "wilaya": "Oran",
                "ai_score": 78,
                "ai_summary": "Développeuse Frontend React junior avec 1,5 an d'expérience. Passionnée par le design UI/UX et la création de sites web interactifs et responsive.",
                "strengths": ["Bonne maîtrise de React & CSS modern", "Sensibilité au design UI/UX", "Jeune et très motivée"],
                "weaknesses": ["Expérience globale limitée", "Manque de compétences en tests automatisés"],
                "pipeline_stage": "Reviewed",
                "notes": "CV intéressant pour un poste junior. À contacter pour une première entrevue.",
                "file_name": "Sara_Haddad_Frontend.docx",
                "file_hash": "mock_hash_sara_haddad_987654",
                "file_content_text": "Sara Haddad. Email: sara.haddad@email.com. Phone: 0770456789. Oran. Frontend React Developer. Experience: Front-end Developer at WebStudio Oran, Intern Web Integrator at OranApp. Education: License in Computer Science from University of Oran. Skills: React, JavaScript, HTML, CSS, Tailwind CSS, Git, Figma."
            },
            {
                "name": "Yacine Bensaoula",
                "phone": "0550 11 22 33",
                "email": "y.bensaoula@email.com",
                "education": ["Licence en Génie Logiciel - Université de Tlemcen (2022)"],
                "experience_years": 4.0,
                "experience_details": [
                    "Développeur Fullstack MERN - Tlemcen Code (2023-Présent)",
                    "Développeur Javascript Freelance (2022-2023)"
                ],
                "skills": ["React", "Node.js", "Express", "MongoDB", "JavaScript", "TypeScript", "Redux", "Docker"],
                "languages": ["Arabic", "French", "English"],
                "wilaya": "Tlemcen",
                "ai_score": 89,
                "ai_summary": "Développeur Full-Stack JavaScript expérimenté (4 ans), basé à Tlemcen. Expertise équilibrée entre le développement d'interfaces React dynamiques et d'APIs Node.js.",
                "strengths": ["Double compétence Frontend/Backend", "Autonomie démontrée en Freelance", "Résident à Tlemcen"],
                "weaknesses": ["Manque d'expérience sur les frameworks Python ou Java"],
                "pipeline_stage": "New",
                "notes": "Profil complet. La polyvalence JavaScript est un atout pour nos projets de startup.",
                "file_name": "Yacine_Bensaoula_Fullstack.pdf",
                "file_hash": "mock_hash_yacine_bensaoula_112233",
                "file_content_text": "Yacine Bensaoula. Email: y.bensaoula@email.com. Phone: 0550112233. Tlemcen. Fullstack Developer with 4 years experience. Experience: Fullstack MERN Developer at Tlemcen Code, Freelance Javascript Developer. Education: License in Software Engineering from University of Tlemcen. Skills: React, Node.js, Express, MongoDB, JavaScript, TypeScript, Redux, Docker."
            },
            {
                "name": "Rania Mansouri",
                "phone": "0662 33 44 55",
                "email": "rania.m@email.com",
                "education": ["Master en Data Science - Université de Constantine (2025)"],
                "experience_years": 1.0,
                "experience_details": [
                    "Data Analyst Intern - Telecom Constantine (2025)",
                    "Projets académiques en Machine Learning"
                ],
                "skills": ["Python", "SQL", "Data Analysis", "Machine Learning", "Tableau", "Git", "Pandas", "Scikit-Learn"],
                "languages": ["Arabic", "French", "English", "German"],
                "wilaya": "Constantine",
                "ai_score": 75,
                "ai_summary": "Data Scientist récemment diplômée (Constantine). Compétences solides en analyse de données et modélisation prédictive avec Python et SQL.",
                "strengths": ["Solides bases en mathématiques & statistiques", "Maîtrise de SQL et des outils d'analyse", "Langues multiples (dont Allemand)"],
                "weaknesses": ["Manque d'expérience professionnelle en entreprise", "Compétences de développement logiciel général à parfaire"],
                "pipeline_stage": "New",
                "notes": "Profil académique brillant. Bon potentiel pour un poste junior en Data.",
                "file_name": "Rania_Mansouri_Data.pdf",
                "file_hash": "mock_hash_rania_mansouri_334455",
                "file_content_text": "Rania Mansouri. Email: rania.m@email.com. Phone: 0662334455. Constantine. Data Scientist. Experience: Data Analyst Intern at Telecom Constantine. Education: Master in Data Science from University of Constantine. Skills: Python, SQL, Data Analysis, Machine Learning, Tableau, Git, Pandas, Scikit-Learn. Languages: Arabic, French, English, German."
            }
        ]
        for c in seed_data:
            try:
                db.add_candidate(c)
            except Exception as e:
                print(f"Error seeding candidate {c['name']}: {e}")

# Serve Single Page Application UI
@app.get("/", response_class=HTMLResponse)
def read_root():
    # Read templates/index.html directly for convenience or use templates engine
    index_path = os.path.join(TEMPLATES_DIR, "index.html")
    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read(), status_code=200)
    return HTMLResponse("<h2>index.html is being created... Please refresh in a moment.</h2>")

# 1. API: Get all candidates
@app.get("/api/candidates")
def get_candidates():
    try:
        return db.get_all_candidates()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 2. API: Get top 10 candidates by AI Score
@app.get("/api/candidates/top10")
def get_top10_candidates():
    try:
        candidates = db.get_all_candidates()
        top10 = sorted(candidates, key=lambda c: c.get("ai_score", 0), reverse=True)[:10]
        return top10
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 3. API: Get single candidate
@app.get("/api/candidates/{candidate_id}")
def get_candidate(candidate_id: int):
    candidate = db.get_candidate(candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return candidate

# 4. API: Delete candidate
@app.delete("/api/candidates/{candidate_id}")
def delete_candidate(candidate_id: int):
    candidate = db.get_candidate(candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    try:
        db.delete_candidate(candidate_id)
        return {"status": "success", "message": "Candidate deleted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 5. API: Update pipeline stage
@app.put("/api/candidates/{candidate_id}/stage")
def update_stage(candidate_id: int, stage: str = Query(...)):
    if stage not in ["New", "Reviewed", "Interview", "Test", "Accepted", "Rejected"]:
        raise HTTPException(status_code=400, detail="Invalid pipeline stage")
    try:
        db.update_candidate_stage(candidate_id, stage)
        return {"status": "success", "message": f"Stage updated to {stage}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 6. API: Update HR notes
@app.put("/api/candidates/{candidate_id}/notes")
def update_notes(candidate_id: int, data: dict):
    notes = data.get("notes", "")
    try:
        db.update_candidate_notes(candidate_id, notes)
        return {"status": "success", "message": "Notes updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 7. API: Smart Search (Fuzzy matching + Skills Boost + Location Penalty + Exp Normalization)
@app.get("/api/search")
def search_candidates(q: str = Query("")):
    try:
        candidates = db.get_all_candidates()
        ranked, relevance_scores = search.rank_candidates(candidates, q)
        
        # Attach relevance score to candidate objects for UI
        results = []
        for cand in ranked:
            cand_copy = dict(cand)
            cand_copy["relevance_score"] = relevance_scores.get(cand["id"], 0.0)
            results.append(cand_copy)
            
        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 8. API: Upload CV (with exact and fuzzy duplicate detection)
@app.post("/api/upload")
async def upload_cv(
    file: UploadFile = File(...),
    force: bool = Form(False)
):
    try:
        filename = file.filename
        file_bytes = await file.read()
        file_hash = parser.get_file_hash(file_bytes)

        # 1. Direct Hash Duplicate Check
        hash_duplicate = db.check_duplicate_by_hash(file_hash)
        if hash_duplicate and not force:
            return {
                "status": "duplicate",
                "duplicate_type": "exact_hash",
                "message": f"Ce CV exact a déjà été téléchargé pour le candidat: {hash_duplicate['name']}.",
                "candidate": hash_duplicate
            }

        # Save to temp file to read text content
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".pdf", ".docx"]:
            raise HTTPException(status_code=400, detail="Seuls les formats PDF et DOCX sont acceptés.")

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as temp:
            temp.write(file_bytes)
            temp_path = temp.name

        # Extract Text
        if ext == ".pdf":
            text = parser.extract_text_from_pdf(temp_path)
        else:
            text = parser.extract_text_from_docx(temp_path)

        # Clean up temp file
        os.remove(temp_path)

        if not text.strip() and not force:
            # If text extraction is completely empty (scanned image or empty document),
            # we will let the parser handle fallback or mock generation, but if it has no text,
            # we warn the user.
            pass

        # 2. Check for Duplicate by Email or Phone
        parsed_mock = parser.parse_resume_text(text, filename)
        email = parsed_mock.get("email")
        phone = parsed_mock.get("phone")
        
        contact_duplicate = db.check_duplicate_by_email_or_phone(email, phone)
        if contact_duplicate and not force:
            return {
                "status": "duplicate",
                "duplicate_type": "contact_match",
                "message": f"Un candidat portant le même Email ou Téléphone existe déjà: {contact_duplicate['name']}.",
                "candidate": contact_duplicate
            }

        # 3. Check for Fuzzy Content Duplicate (Similarity)
        all_candidates = db.get_all_candidates()
        highest_similarity = 0.0
        similar_candidate = None
        
        for cand in all_candidates:
            sim = parser.calculate_jaccard_similarity(text, cand.get("file_content_text", ""))
            if sim > highest_similarity:
                highest_similarity = sim
                similar_candidate = cand

        # Threshold of 85% similarity indicates a fuzzy content duplicate
        if highest_similarity >= 0.85 and not force:
            return {
                "status": "duplicate",
                "duplicate_type": "content_similarity",
                "similarity_score": round(highest_similarity * 100, 1),
                "message": f"هذا الملف مكرر بنسبة {round(highest_similarity * 100)}% مع ملف المترشح {similar_candidate['name']}.",
                "candidate": {"id": similar_candidate["id"], "name": similar_candidate["name"]}
            }

        # If duplicate check passed (or force is True), save candidate
        # If force is true and we found duplicate, let's delete the old one or just save as new
        if force:
            # If we are overwriting, we can delete the duplicate to keep the database clean
            if hash_duplicate:
                db.delete_candidate(hash_duplicate["id"])
            elif contact_duplicate:
                db.delete_candidate(contact_duplicate["id"])
            elif similar_candidate and highest_similarity >= 0.85:
                db.delete_candidate(similar_candidate["id"])

        # Insert candidate
        parsed_mock["file_hash"] = file_hash
        new_id = db.add_candidate(parsed_mock)
        
        # Save actual file to uploads folder for preview simulations
        saved_file_name = f"{new_id}_{filename}"
        saved_file_path = os.path.join(UPLOAD_DIR, saved_file_name)
        with open(saved_file_path, "wb") as f:
            f.write(file_bytes)

        return {
            "status": "success",
            "message": "CV téléchargé et analysé avec succès !",
            "candidate_id": new_id
        }

    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# 9. API: QR Code Generator (Returns inline Svg XML)
@app.get("/api/candidates/{candidate_id}/qr")
def get_candidate_qr(candidate_id: int):
    candidate = db.get_candidate(candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
        
    try:
        # We generate a QR code linking to the candidate profile page
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
            image_factory=qrcode.image.svg.SvgPathImage
        )
        # Deep link inside our SPA (will resolve via Hash router or static route)
        # Using a general local link, or host-independent link
        profile_url = f"http://localhost:8000/#profile/{candidate_id}"
        qr.add_data(profile_url)
        qr.make(fit=True)
        
        img = qr.make_image()
        
        # Convert to string (SVG XML)
        svg_xml = img.to_string(encoding="unicode")
        
        return Response(content=svg_xml, media_type="image/svg+xml")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 10. API: Export Excel
@app.get("/api/export/excel")
def export_excel(q: str = Query("")):
    try:
        # Get candidates (filtered if query is present)
        candidates = db.get_all_candidates()
        if q:
            candidates, _ = search.rank_candidates(candidates, q)

        # Create OpenPyXL workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Météo des Mutiles" if q else "Candidats TalentFlow"
        
        # Set sheet directions
        ws.views.sheetView[0].showGridLines = True
        
        # Colors & styling
        navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
        data_font = Font(name="Segoe UI", size=10)
        bold_data_font = Font(name="Segoe UI", size=10, bold=True)
        
        thin_side = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Add Title Row
        ws.merge_cells("A1:I1")
        ws["A1"] = "Rapport des Candidats - TalentFlow AI" if not q else f"Résultats de recherche pour: '{q}'"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Add headers
        headers = ["ID", "Nom", "Téléphone", "Email", "Wilaya", "Expérience (Ans)", "AI Score", "Compétences", "Statut Etape"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
        ws.row_dimensions[3].height = 25

        # Add data
        for row_idx, cand in enumerate(candidates, 4):
            ws.cell(row=row_idx, column=1, value=cand["id"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=cand["name"]).font = bold_data_font
            ws.cell(row=row_idx, column=3, value=cand["phone"])
            ws.cell(row=row_idx, column=4, value=cand["email"])
            ws.cell(row=row_idx, column=5, value=cand["wilaya"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=6, value=cand["experience_years"]).alignment = Alignment(horizontal="center")
            
            # AI Score with color conditional styling
            score_cell = ws.cell(row=row_idx, column=7, value=f"{cand['ai_score']}%")
            score_cell.alignment = Alignment(horizontal="center")
            score_cell.font = bold_data_font
            if cand['ai_score'] >= 85:
                score_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # green
            elif cand['ai_score'] >= 70:
                score_cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # yellow
            else:
                score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # red

            ws.cell(row=row_idx, column=8, value=cand["skills"])
            ws.cell(row=row_idx, column=9, value=cand["pipeline_stage"]).alignment = Alignment(horizontal="center")
            
            # Apply borders and default fonts to all cells in the row
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.font or cell.font.name != "Segoe UI":
                    cell.font = data_font
                cell.border = border_all
                
            ws.row_dimensions[row_idx].height = 20

        # Adjust columns width
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            # Skip checking the title row A1 length to avoid super wide columns
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Write to memory stream
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="talentflow_candidates.xlsx"'
        }
        return Response(content=stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    # Start server
    uvicorn.run(app, host="127.0.0.1", port=8000)
